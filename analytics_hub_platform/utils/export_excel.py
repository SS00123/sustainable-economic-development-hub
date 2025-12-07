"""
Excel Export Utility
Sustainable Economic Development Analytics Hub
Ministry of Economy and Planning

Generates professional Excel workbooks with formatted data and charts.
"""

from io import BytesIO
from datetime import datetime
from typing import Optional, List, Dict
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
    )
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from analytics_hub_platform.config.branding import BRANDING
from analytics_hub_platform.config.theme import get_theme


def generate_excel_workbook(
    data: pd.DataFrame,
    title: str = "Sustainability Analytics Data",
    summary_metrics: Optional[dict] = None,
    include_charts: bool = True,
    language: str = "en",
) -> BytesIO:
    """
    Generate an Excel workbook from indicator data.
    
    Args:
        data: DataFrame containing indicator data
        title: Workbook title
        summary_metrics: Optional dict of key metrics for summary sheet
        include_charts: Whether to include chart sheets
        language: Language code (en/ar)
    
    Returns:
        BytesIO buffer containing the Excel file
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install with: pip install openpyxl"
        )
    
    buffer = BytesIO()
    theme = get_theme()
    
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Define styles
    header_fill = PatternFill(
        start_color=theme.colors.primary.lstrip("#"),
        end_color=theme.colors.primary.lstrip("#"),
        fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    alt_fill = PatternFill(
        start_color="F8FAFC",
        end_color="F8FAFC",
        fill_type="solid"
    )
    
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    
    title_font = Font(bold=True, size=16, color=theme.colors.primary.lstrip("#"))
    subtitle_font = Font(size=11, color="64748B")
    
    # Summary Sheet
    ws_summary["A1"] = title
    ws_summary["A1"].font = title_font
    ws_summary.merge_cells("A1:D1")
    
    ws_summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_summary["A2"].font = subtitle_font
    
    ws_summary["A3"] = f"Prepared by: {BRANDING['author_name']}"
    ws_summary["A3"].font = subtitle_font
    
    # Summary metrics
    if summary_metrics:
        ws_summary["A5"] = "Key Metrics"
        ws_summary["A5"].font = Font(bold=True, size=12)
        
        row = 6
        for metric_name, metric_value in summary_metrics.items():
            ws_summary.cell(row=row, column=1, value=metric_name)
            ws_summary.cell(row=row, column=2, value=str(metric_value))
            ws_summary.cell(row=row, column=1).border = thin_border
            ws_summary.cell(row=row, column=2).border = thin_border
            row += 1
        
        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 20
    
    # Data Sheet
    ws_data = wb.create_sheet("Data")
    
    if not data.empty:
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
                
                # Header row styling
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                else:
                    # Alternate row colors
                    if r_idx % 2 == 0:
                        cell.fill = alt_fill
                    cell.border = thin_border
        
        # Auto-adjust column widths
        for col_idx, column in enumerate(data.columns, 1):
            max_length = max(
                len(str(column)),
                data[column].astype(str).str.len().max() if len(data) > 0 else 0
            )
            adjusted_width = min(max_length + 2, 40)
            ws_data.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Freeze header row
        ws_data.freeze_panes = "A2"
    
    # Regional Summary Sheet
    if not data.empty and "region" in data.columns:
        ws_regional = wb.create_sheet("Regional Summary")
        
        regional_summary = data.groupby("region").agg({
            col: "mean" for col in data.select_dtypes(include=["float64", "int64"]).columns
            if col not in ["year", "quarter", "tenant_id"]
        }).round(2)
        
        ws_regional["A1"] = "Regional Performance Summary"
        ws_regional["A1"].font = title_font
        ws_regional.merge_cells("A1:E1")
        
        for r_idx, row in enumerate(dataframe_to_rows(regional_summary, index=True, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws_regional.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == 3:
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    cell.border = thin_border
                    if r_idx % 2 == 0:
                        cell.fill = alt_fill
    
    # Trend Sheet with Chart
    if include_charts and not data.empty and "year" in data.columns:
        ws_trend = wb.create_sheet("Trends")
        
        # Prepare trend data
        if "sustainability_index" in data.columns:
            trend_data = data.groupby("year")["sustainability_index"].mean().reset_index()
            
            ws_trend["A1"] = "Sustainability Index Trend"
            ws_trend["A1"].font = title_font
            
            ws_trend["A3"] = "Year"
            ws_trend["B3"] = "Sustainability Index"
            ws_trend["A3"].fill = header_fill
            ws_trend["A3"].font = header_font
            ws_trend["B3"].fill = header_fill
            ws_trend["B3"].font = header_font
            
            for idx, row in trend_data.iterrows():
                ws_trend.cell(row=4 + idx, column=1, value=int(row["year"]))
                ws_trend.cell(row=4 + idx, column=2, value=round(row["sustainability_index"], 2))
            
            # Add line chart
            if len(trend_data) >= 2:
                chart = LineChart()
                chart.title = "Sustainability Index Over Time"
                chart.style = 10
                chart.x_axis.title = "Year"
                chart.y_axis.title = "Index Value"
                chart.width = 15
                chart.height = 8
                
                data_ref = Reference(ws_trend, min_col=2, min_row=3, max_row=3 + len(trend_data))
                cats_ref = Reference(ws_trend, min_col=1, min_row=4, max_row=3 + len(trend_data))
                
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                
                ws_trend.add_chart(chart, "D3")
    
    # Metadata Sheet
    ws_meta = wb.create_sheet("Metadata")
    
    meta_data = [
        ("Report Title", title),
        ("Generation Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Data Rows", len(data) if not data.empty else 0),
        ("Prepared By", BRANDING["author_name"]),
        ("Contact Email", BRANDING["author_email"]),
        ("Contact Phone", BRANDING["author_mobile"]),
        ("Platform", "Sustainable Economic Development Analytics Hub"),
        ("Organization", "Ministry of Economy and Planning"),
    ]
    
    ws_meta["A1"] = "Export Metadata"
    ws_meta["A1"].font = title_font
    
    for idx, (key, value) in enumerate(meta_data, 3):
        ws_meta.cell(row=idx, column=1, value=key).font = Font(bold=True)
        ws_meta.cell(row=idx, column=2, value=value)
    
    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 40
    
    # Save workbook
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def generate_simple_excel(
    data: pd.DataFrame,
    sheet_name: str = "Data",
) -> BytesIO:
    """
    Generate a simple Excel file from a DataFrame.
    
    Args:
        data: DataFrame to export
        sheet_name: Name of the worksheet
    
    Returns:
        BytesIO buffer containing the Excel file
    """
    buffer = BytesIO()
    
    if OPENPYXL_AVAILABLE:
        data.to_excel(buffer, sheet_name=sheet_name, index=False, engine="openpyxl")
    else:
        # Fallback to CSV in Excel format
        data.to_csv(buffer, index=False)
    
    buffer.seek(0)
    return buffer


def generate_multi_sheet_excel(
    sheets: Dict[str, pd.DataFrame],
    title: str = "Analytics Report",
) -> BytesIO:
    """
    Generate an Excel file with multiple sheets.
    
    Args:
        sheets: Dictionary mapping sheet names to DataFrames
        title: Overall report title
    
    Returns:
        BytesIO buffer containing the Excel file
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl required for Excel export")
    
    buffer = BytesIO()
    theme = get_theme()
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    header_fill = PatternFill(
        start_color=theme.colors.primary.lstrip("#"),
        end_color=theme.colors.primary.lstrip("#"),
        fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    
    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])  # Excel sheet name limit
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
        
        # Auto-adjust widths
        for col_idx, column in enumerate(df.columns, 1):
            max_length = max(len(str(column)), 10)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
